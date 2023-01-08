import pandas as pd
import numpy as np
import os
import openpyxl
import re
# Method that checks if a data point is null
# Parameter: string
# Returns: yes if string is null, no if not null
def isNaN(string):
    return string != string

# Method that returns the index of the last occurence of a certain value in a string
# Parameters: string, value to look for
# Returns: the index and the count of that value
def index(string, val):
    count = 0
    dash_index = 0
    for i, n in enumerate(string):
        if n == val:
            count += 1
            dash_index = i
    return dash_index, count

# Takes in 2 characters
# Returns if they are both capital letters (A - Z) or digits (1 - 10)
def are_same_type(char1, char2):
    return ((58 > ord(char1) > 47) and (58 > ord(char2) > 47)) or ((91 > ord(char1) > 64) and (91 > ord(char2) > 64)) 


# Used to expand a Cell Marque ID from abbreviated form. 
# Uses string manipulation
# Inputs one ID as a string and returns a list of the new IDs
# AA-1A-G -> AA-1A, AA-1B, AA-1C...AA-1G
def expand(ID):
    newIDS = np.array([])
    second_dash_index, dash_count = index(ID, "-")  
    if len(ID) > 1 and dash_count > 0:
        first_letter_index = second_dash_index - 1   
        first_letter = ID[first_letter_index]
        first_letter_ASCII = ord(first_letter)  
        if (not isNaN(ID[second_dash_index + 1])):
            final_letter = ID[second_dash_index + 1]
            final_letter_ASCII = ord(final_letter)
            
        if dash_count > 1 and (are_same_type(first_letter, final_letter)):        
            newIDS = np.append(newIDS, ID[0 : second_dash_index])
            for i in range (final_letter_ASCII - first_letter_ASCII):
                newID = ID[0 : first_letter_index] + chr(first_letter_ASCII + 1 + i)
                newIDS = np.append(newIDS, newID)
            dash_count = 0
        else: 
            newIDS = np.append(newIDS, ID)
    return newIDS
    
# Splits up lists of Block ID on spaces, commas, amprisands, or semicolons
# Catches a few cases that are formatted differently (with spaces in their abbreviated form)
# Removes extra spaces 
# Then expands each block ID set by calling the expand method
# Returns the new, expanded IDs for an entire row
def generate_new_IDS(data, old_CM_IDS):
    new_CM_IDS = np.array([])
    if not isNaN(old_CM_IDS):
        if not (old_CM_IDS.startswith("BRN CRB") or old_CM_IDS.startswith("BRN CEBL") or old_CM_IDS.startswith("CIN I") or old_CM_IDS.startswith("CIN II") or old_CM_IDS.startswith("CIN III") or old_CM_IDS.startswith("Umbilical Cord") or old_CM_IDS.startswith("Spermatic Cord")):
            old_CM_IDS = re.split(', | | & |; ', old_CM_IDS)  
        else: 
            old_CM_IDS = re.split(', | & |; ', old_CM_IDS)          
        for ID in old_CM_IDS:
            ID = ID.strip()
            if (ID == ' ') or (ID == '&') or (ID == '') or len(ID) < 2:
                old_CM_IDS.remove(ID)        
        for ID in old_CM_IDS: 
            expanded_IDS = expand(ID)
            new_CM_IDS = np.append(new_CM_IDS, expanded_IDS)
    return new_CM_IDS

# Adds the new rows for expanded block IDs
def add_new_rows_for_IDS(data, i):
    old_CM_IDS = data.loc[i, "Block ID #"]
    new_CM_IDS = generate_new_IDS(data, old_CM_IDS)
    new_rows = pd.DataFrame(columns = data.columns)
    if new_CM_IDS.size > 1:
        data.at[i, "Block ID #"] = new_CM_IDS[0]
        for j in range(new_CM_IDS.size - 1):
            newID_row = pd.DataFrame(data = data.iloc[[i]], columns = data.columns)
            newID_row.at[i, "Block ID #"] = new_CM_IDS[j + 1]
            newID_row.at[i, "Total # Blocks"] = 1
            newID_row.at[i, "# Blocks used"] = 1
            new_rows = pd.concat([new_rows, newID_row], ignore_index = True)
        data = pd.concat([data, new_rows], ignore_index = True)
    data.reset_index(drop = True, inplace = True)
    return data

# Import Data
pwd = "C:/Users/x248189/OneDrive - MerckGroup/Documents"
path = "//Frozen Closet Blocks Highlighted Edit.xlsx"
original_closet_data = openpyxl.load_workbook(pwd + path)

# Remove the unnecessary sheets
closet_sheetnames = original_closet_data.sheetnames
closet_sheetnames.remove("METRICS")
closet_sheetnames.remove("delete")

# Setup Empty DataFrame with Columns
closet_blocks = pd.read_excel(pwd + path, sheet_name = "ADRENAL")
closet_blocks.columns = [j.strip() for j in closet_blocks.columns]
# Source - vendor that gave us the block 
list(closet_blocks.columns).append("Source")
# Category - the sheet name that the block is located in
list(closet_blocks.columns).append("Category")
closet_columns = closet_blocks.columns

# Empty closet blocks data frame to add to
closet_blocks = closet_blocks.iloc[0:0]

k = 0
# Creating one sheet for all Closet Blocks, retaining the info of Category (sheet name) so no information is lost.
for sheet in closet_sheetnames:
    k+=1
    print(k)
    print(sheet)  
    closet_sheet = pd.read_excel(pwd + path, sheet_name = sheet)
    # Removing unnamed, hidden columns (caused error)
    closet_sheet = closet_sheet.loc[:, ~closet_sheet.columns.str.contains('^Unnamed')]
    closet_sheet.columns = closet_columns
    # drop completely blank rows
    closet_sheet = closet_sheet.dropna(how = 'all') 
    # reset indexing
    closet_sheet.reset_index(inplace = True, drop = True)
    num_rows = closet_sheet.shape[0]
    closet_sheet = closet_sheet.assign(Category = sheet)
    # add this sheet in to the closet blocks data frame
    closet_blocks = pd.concat([closet_blocks, closet_sheet], ignore_index = True)

# Export Intermittently as a QC
# So far, we have just combined all the closet blocks together into one sheet.
closet_blocks.to_excel(pwd + "//Output Files/Simply Combined Closet Blocks.xlsx", index = False)

# We have to now separate each row so that it represents only one block.
closet_blocks = pd.read_excel(pwd + "//Output Files/Simply Combined Closet Blocks.xlsx")
closet_blocks['# Blocks used'] = closet_blocks['# Blocks used'].fillna(0)
closet_blocks = closet_blocks.dropna(subset = 'Total # Blocks')
closet_blocks.reset_index(inplace = True, drop = True)
source_data = np.array([])
for j in range(closet_blocks.shape[0]):
    print(j)
    surgical_number = str(closet_blocks.loc[j, "Surgical Number"])
    if surgical_number.lower().strip().startswith('p') or surgical_number.lower().strip().startswith('d'):
        source = "Princeton Baptist Medical Center: 701 Princeton Ave. S.W.. Birmingham AL 35211"
    elif 'lv' in surgical_number.lower():
        source = "Lakeview Hospital: 630 East Medical Dr., Bountiful UT 84010"
    elif 'ml' in surgical_number.lower() or 'mf' in surgical_number.lower():
        source = "NatPkMedCtr AMI HotSprings"
    else:
        source = "UNKNOWN"
    source_data = np.append(source_data, source)
closet_blocks = closet_blocks.assign(Source = source_data)

for i in range(closet_blocks.shape[0]):
    print(i)
    if not str(closet_blocks.loc[i, "Total # Blocks"]).isdigit():
        closet_blocks.at[i, "Total # Blocks"] = 0
    if not str(closet_blocks.loc[i, "# Blocks used"]).isdigit():
        closet_blocks.at[i, "# Blocks used"] = 0
    total_blocks_for_row = 0
    blocks_used_for_row = 0
    total_blocks_for_row = int(closet_blocks.loc[i, "Total # Blocks"])
    blocks_used_for_row = int(closet_blocks.loc[i, "# Blocks used"])
    # if not already in the format we want
    if (total_blocks_for_row > 1 and total_blocks_for_row > blocks_used_for_row):
        unqualified_number_for_row = total_blocks_for_row - blocks_used_for_row
        unqualified_rows = pd.DataFrame(columns = closet_columns)
        if blocks_used_for_row > 0:
            for j in range(unqualified_number_for_row):
                new_row = pd.DataFrame(data = closet_blocks.iloc[[i]], columns = closet_columns)
                new_row.at[i, "Total # Blocks"] = 1
                new_row.at[i, "# Blocks used"] = 0
                new_row.at[i, "CM Number"] = ""
                new_row.at[i, "Initials/H&E Date"] = ""
                unqualified_rows = pd.concat([unqualified_rows, new_row], ignore_index = True)
            closet_blocks.at[i, "Total # Blocks"] = blocks_used_for_row
        else: 
            for j in range(unqualified_number_for_row - 1):
                new_row = pd.DataFrame(data = closet_blocks.iloc[[i]], columns = closet_columns)
                new_row.at[i, "Total # Blocks"] = 1
                new_row.at[i, "# Blocks used"] = 0
                new_row.at[i, "CM Number"] = ""
                new_row.at[i, "Initials/H&E Date"] = ""
                unqualified_rows = pd.concat([unqualified_rows, new_row], ignore_index = True)
            closet_blocks.at[i, "Total # Blocks"] = 1
        closet_blocks = pd.concat([closet_blocks, unqualified_rows], ignore_index = True)

closet_blocks.reset_index(inplace = True, drop = True)

print("There are a total of {} non-blank rows in Closet Blocks (Expanded Unqualified)" .format(closet_blocks.shape[0]))
closet_blocks.to_excel(pwd + "//Output Files/Expanded and Combined Closet Blocks.xlsx", index = False)

closet_blocks.rename(columns = {'CM Number':'Block ID #'}, inplace = True)
unqualified_closet_blocks = pd.DataFrame(columns = closet_columns)

# put unqualified rows in a separate sheet
error_rows = closet_blocks.loc[closet_blocks["Total # Blocks"] < closet_blocks["# Blocks used"]]
error_rows.reset_index(inplace = True, drop = True)

unqualified_closet_blocks = closet_blocks.loc[closet_blocks["# Blocks used"] == 0]
unqualified_closet_blocks.reset_index(inplace = True, drop = True)

cleaned_closet_blocks = closet_blocks.loc[closet_blocks["# Blocks used"] != 0]
cleaned_closet_blocks.reset_index(inplace = True, drop = True)

for i in range(cleaned_closet_blocks.shape[0]):
    cleaned_closet_blocks = add_new_rows_for_IDS(cleaned_closet_blocks, i)
    # We have made it so each row represents only one block
    # And they are all qualified blocks
    # So 1 block total, 1 block used
    cleaned_closet_blocks.at[i, "Total # Blocks"] = 1
    cleaned_closet_blocks.at[i, "# Blocks used"] = 1
cleaned_closet_blocks.reset_index(inplace = True, drop = True)

print("There are a total of {} non-blank rows in Closet Blocks (Cleaned)" .format(cleaned_closet_blocks.shape[0]))
cleaned_closet_blocks.to_excel(pwd + "//Output Files/Cleaned Closet Blocks.xlsx", index = False)

print("There are a total of {} non-blank rows in Closet Blocks (Unqualified)" .format(unqualified_closet_blocks.shape[0]))
unqualified_closet_blocks.to_excel(pwd + "//Output Files/Unqualified Closet Blocks.xlsx", index = False)

print("There are a total of {} error rows in Closet Blocks" .format(error_rows.shape[0]))
error_rows.to_excel(pwd + "//Output Files/Error Rows.xlsx", index = False)