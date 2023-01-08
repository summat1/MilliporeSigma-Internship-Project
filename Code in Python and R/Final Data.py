# Import Packages
import pandas as pd
import numpy as np
import os
import openpyxl
import re
import datetime

# Returns true if the string is NA
def isNaN(string):
    return string != string

# Takes in a datetime.datetime object and returns it in the specificed format
# 01 Jan 2000
def convert(time):
    month_name = time.strftime("%b").upper()
    new_time = (str(time.day) + " " + str(month_name) + " " + str(time.year))
    return new_time

# Takes in a date and returns how many digits there are following the last slash
# Helps determine if the whole year is there - MM/DD/YYYY, or not - MM/DD/YY
def dateType(date):
    date = date.strip()
    index = 0
    i = 0
    for s in date: 
        i += 1
        if s == "/":
            index = i
    return len(date) - index

# Checks how many slashes there are in the string
def slashCheck(date):
    i = 0
    for s in str(date):
        if s == "/":
            i +=1
    return i == 2

# Removes parantheses via regular expressions
def removeParan(date):
    date = re.sub(r'\(.*?\)','', date)
    return date

# A robust algorithm to take any CM ID # and split into its 4 components
# as outlined in the tissue data syntax guide.
# The comments will walk through an example to explain what happens at which point
# input: Block ID#
# output: list of 4 components
# Example: AA-10A1
def splitID(block_id):
    block_id = block_id.strip()
    # [AA, 10A1] 
    dash_split = block_id.split("-") 
    total_split = len(dash_split)
    # This branch is if the tissue type itself has dashes in it.
    # Example: "STM-CA-DIF"
    if total_split != 2:
        tissue_type = dash_split[0]
        for i in range(total_split - 2):
            tissue_type = tissue_type + "-"+ dash_split[i+1]
    else:
        # AA 
        tissue_type = dash_split[0]
    # 10A1
    patient_and_sample = dash_split[-1]
    subsplit_number = ""
    for i in range(len(patient_and_sample)):
        if patient_and_sample[i].isalpha():
            break
    #10
    patient_number = patient_and_sample[0:i]
    #A1
    sample_number = patient_and_sample[i:]
    for j in range(len(sample_number)):
        if not sample_number[j].isalpha():
            break
    sample = sample_number
    # A
    sample_number = sample_number[0:j]
    # 1
    subsplit_number = sample[j:]
    return tissue_type, patient_number, sample_number, subsplit_number


# Initializing Dataframes of Formatted Data and QTI
pwd = "C:/Users/x248189/OneDrive - MerckGroup/Documents"
path = "/Output Files//QTI + Source Data Merged.xlsx"
data = pd.read_excel(pwd + path)

path = "/Output Files//Cleaned Qualified Tissue Inventory.xlsx"
active_inventory = pd.read_excel(pwd + path)

# Finding each block that wasn't matched
active_blocks = active_inventory['Block ID #']
found_blocks = data['Block ID #']
not_matched_indices = active_blocks[~active_blocks.isin(found_blocks)]
mask = active_inventory['Block ID #'].isin(not_matched_indices)
not_matched_rows = active_inventory[mask]
not_matched_rows.dropna(subset = "Block ID #", inplace = True)

# Renaming "Initials" so there is no duplicate column names
data.columns.values[11] = "Rename Initials"
not_matched_rows.columns.values[11] = "Rename Initials"

# Changing index to allow for concatentation of the not matched rows with matched data
data.set_index('Block ID #', inplace = True)
not_matched_rows.set_index('Block ID #', inplace = True)

# Concatentating to recreate full QTI
data = pd.concat([data, not_matched_rows])
data.sort_values(by=['Block ID #'], inplace=True)
data.rename(columns = {"Date" : "Date Changed"}, inplace = True)
data = data[~data.index.duplicated(keep='first')]

dx_change = np.array([])
source_type = np.array([])

# fixing a certain case
data.at["STM-CA-DIF-18A1", "Date Added"] = "11/21/2013"

# Iterating through the entire sheet
for i, row in data.iterrows():
    # Changing format of date added
    date_added = data.loc[i, "Date Added"]
    if type(date_added) == datetime.datetime:
        data.at[i, "Date Added"] = convert(date_added)
    elif not isNaN(date_added) and slashCheck(date_added):
        date_added = date_added.strip()
        if dateType(date_added) < 3: 
            date_added = datetime.datetime.strptime(date_added, '%m/%d/%y')
        else: 
            date_added = datetime.datetime.strptime(date_added, '%m/%d/%Y')
        data.at[i, "Date Added"] = convert(date_added)

    # Changing format of date changed
    date_changed = data.loc[i, "Date Changed"]
    if type(date_changed) == datetime.datetime:
        data.at[i, "Date Changed"] = convert(date_changed)
    elif not isNaN(date_changed):       
        if slashCheck(date_changed):
            date_changed = removeParan(date_changed)
            date_changed = date_changed.strip()
            if dateType(date_changed) < 3: 
                date_changed = datetime.datetime.strptime(date_changed, '%m/%d/%y')
            else: 
                date_changed = datetime.datetime.strptime(date_changed, '%m/%d/%Y')
            data.at[i, "Date Changed"] = convert(date_changed)

    # Changing format of date received
    date_received = data.loc[i, 'Receiving Date']
    if type(date_received) == datetime.datetime:
        data.at[i, "Receiving Date"] = convert(date_received)
    elif not isNaN(date_received):
        date_received = removeParan(str(date_received))
        date_received = date_received.strip()
        if slashCheck(date_received):
            if dateType(date_received) < 3: 
                date_received = datetime.datetime.strptime(date_received, '%m/%d/%y')
            else: 
                date_received = datetime.datetime.strptime(date_received, '%m/%d/%Y')
            data.at[i, "Receiving Date"] = convert(date_received)

    # Adding Dx Change column
    if isNaN(data.loc[i, "Previous Block ID"]):
        dx_change = np.append(dx_change, "No")
    else:
        dx_change = np.append(dx_change, "Yes")

    # Identifying source type depending on which rows are filled in from closet/customer block data
    if data.loc[i, ["Receiving Date", "Demographic Info"]].isna().all() and data.loc[i, ["Surgical Number", "# Blocks Discarded"]].isna().all():
        source_type = np.append(source_type, "")
    elif data.loc[i, ["Receiving Date", "Demographic Info"]].isna().all():
        source_type = np.append(source_type, "Closet Blocks")
    elif data.loc[i, ["Surgical Number", "# Blocks Discarded"]].isna().all():
        source_type = np.append(source_type, "Customer Blocks")

data.insert(8, "Dx Change? Yes/No", dx_change)
source = data.pop("Source")
data.insert(12, "Source", source)
data.insert(13, "Source Type", source_type)

# Append source data to child blocks
block_IDS = list(data.index.values)
for i in range(len(block_IDS)-1):
    block_ID = block_IDS[i]
    id_components = splitID(block_ID)
    tissue_and_patient_id = id_components[0] + id_components[1]
    next_block_ID = block_IDS[i+1]
    id_components_next = splitID(next_block_ID)
    tissue_and_patient_id_next = id_components_next[0] + id_components_next[1]
    if isNaN(data.loc[next_block_ID, "Source"]) and (tissue_and_patient_id == tissue_and_patient_id_next) and not id_components_next[3] == "":
        data.at[next_block_ID, "Source"] = data.at[block_ID, "Source"]
        data.at[next_block_ID, "Source Type"] = "Sub-split Block"

no_source = data[data["Source"].isna()]
no_source_blocks = no_source.index.values
no_source_blocks = pd.DataFrame(no_source_blocks)
no_source_blocks.to_excel(pwd + "//Output Files/No Source.xlsx")
# Positive Antibody Scores
pos = pd.DataFrame(columns = ["Block ID", "Test Number", "Antibody", "Score", "Clone", "Staining Pattern", "Percentage of Cells Stained"])
positive_antibody_score = data.loc["COL-102A1", "Positive Antibody Scores"]
positive_scores = positive_antibody_score.split(",")
test_number = 0
for test in positive_scores:
    block_ID = "COL-102A1"
    count = 0
    for i in test: 
        count += 1
        if i == ":":
            first = count
            antibody = test[0:count-1].strip()        
        if i == "(":
            second = count
            score = test[first:second-1].strip()
            clone = test[second-1:len(test)]
    for i in range (len(score)):
        if score[i] == 'F':
            score = score[0:i]
            pos.at[test_number, "Clone Staining Pattern"] = "Focal"
        if score[i] == 'N':
            score = score[0:i]
            pos.at[test_number, "Clone Staining Pattern"] = "N"
    pos.at[test_number, "Block ID"] = block_ID
    pos.at[test_number, "Test Number"] = test_number
    pos.at[test_number, "Antibody"] = antibody
    pos.at[test_number, "Score"] = score
    pos.at[test_number, "Clone"] = clone
    test_number += 1
#print("Antibody: {}, Score: {}, Clone: {}".format(antibody, score, clone))
print(pos)
pos.to_excel(pwd + "/Output Files//Positive Scores.xlsx", index = False)

num_identified_sources = len(data[~data.Source.isna()])
print("Out of {} Blocks, we have identified a source for {} of them. This comes out to {:.2f}%.".format(len(block_IDS), num_identified_sources, 100 * (num_identified_sources/len(block_IDS))))
'''
# Validating Closet Blocks
path = "//Frozen Closet Blocks Highlighted Edit.xlsx"
original_closet = openpyxl.load_workbook(pwd + path)
closet_blocks = pd.read_excel(pwd + path, sheet_name = "ADRENAL")
closet_blocks.columns = [j.strip() for j in closet_blocks.columns]
closet_blocks = closet_blocks.iloc[0:0]

closet_sheetnames = original_closet.sheetnames
closet_sheetnames.remove("METRICS")
closet_sheetnames.remove("delete")

total_closet_blocks = 0
for sheet in closet_sheetnames:
    closet_sheet = pd.read_excel(pwd + path, sheet_name = sheet)
    # Removing unnamed columns that were causing error
    closet_sheet = closet_sheet.loc[:, ~closet_sheet.columns.str.contains('^Unnamed')]
    # drop blank rows
    closet_sheet = closet_sheet.dropna(how = 'all') 
    total_closet_blocks += closet_sheet.shape[0]

print("There are {} rows of data in Closet Blocks".format(total_closet_blocks))
path = "/Output Files//Cleaned Closet Blocks.xlsx"
new_closets = pd.read_excel(pwd + path)
print("There are {} qualified blocks that were found in Closet Blocks".format(new_closets.shape[0]))
path = "/Output Files//Unqualified Closet Blocks.xlsx"
unqualified_closets = pd.read_excel(pwd + path)
print("There are {} unqualified blocks that were found in Closet Blocks".format(unqualified_closets.shape[0]))
print("In total, we started with {} rows in Closet Blocks and found {} rows.".format(total_closet_blocks, unqualified_closets.shape[0]+new_closets.shape[0]))

# Validating Customer Blocks
path = "//Frozen Customer Blocks Highlighted Edit.xlsx"
customer_blocks = pd.read_excel(pwd + path, sheet_name = "Affiliated Dermatology", header = None)
original_data = openpyxl.load_workbook(pwd + path)
sheets_to_remove = ["Sheet1", "Sheet2", "Sheet3", "Unknown", "Sheet4", "1", "2", "3", "Sheet5"]
modified_list = original_data.sheetnames
for sheet in sheets_to_remove:
   modified_list.remove(sheet)
total_customer_blocks = 0
for sheet in modified_list:
    data = pd.read_excel(pwd + path, sheet_name = sheet, header = None)
    source = data.iloc[0, 0]
    data = data.drop([0])
    # drops NA columns
    data = data.dropna(how='all')
    data = data.iloc[1:]
    # resets indexing
    data.reset_index(inplace = True, drop = True)
    total_customer_blocks += data.shape[0]
print("There are {} rows of data in Customer Blocks".format(total_customer_blocks))
path = "/Output Files//Cleaned Customer Blocks.xlsx"
new_customers = pd.read_excel(pwd + path)
print("There are {} qualified blocks that were found in Customer Blocks".format(new_customers.shape[0]))
path = "/Output Files//Unqualified Customer Blocks.xlsx"
unqualified_customers = pd.read_excel(pwd + path)
print("There are {} unqualified blocks that were found in Customer Blocks".format(unqualified_customers.shape[0]))
print("In total, we started with {} rows in Customer Blocks and found {} rows.".format(total_customer_blocks, unqualified_customers.shape[0]+new_customers.shape[0]))

'''


data.to_excel(pwd + "/Output Files//Master File of All Data.xlsx")
print("Export Successful.")